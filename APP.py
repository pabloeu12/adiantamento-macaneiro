import streamlit as st
import pandas as pd
import altair as alt
import io
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# CONFIGURAÇÃO DA PÁGINA E CORES
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Conferência Adiantamento Salarial", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ════════════════════════════════════════════════════════════
# BUSCA INTELIGENTE DE LOGOS (Resolve os erros de JPG/PNG/Espaços)
# ════════════════════════════════════════════════════════════
def buscar_logo(palavra_chave):
    """
    Varre a pasta atual e procura qualquer arquivo de imagem que 
    contenha a palavra-chave (ex: 'bwise' ou 'macaneiro').
    """
    try:
        arquivos = os.listdir('.')
        for arq in arquivos:
            if arq.lower().endswith(('.png', '.jpg', '.jpeg')) and palavra_chave in arq.lower():
                return arq
    except Exception:
        pass
    return None

logo_bwise = buscar_logo("bwise")
logo_macaneiro = buscar_logo("macaneiro")

# Estilos CSS Customizados para o Título Centralizado
st.markdown(
    """
    <style>
    .titulo-sistema {
        text-align: center;
        color: #1E3A8A;
        font-family: 'Arial', sans-serif;
        font-size: 1.8rem; /* Reduzido em 40% para melhor equilíbrio visual */
        font-weight: bold;
        line-height: 1.2;
        margin-top: 15px;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# -----------------------------------------------------------------------------
# CABEÇALHO COM LOGOS (Ajustado)
# -----------------------------------------------------------------------------
col1, col2, col3 = st.columns([1, 2.5, 1])

with col1:
    if logo_bwise:
        st.image(logo_bwise, use_column_width=True)
    else:
        st.warning("Logo Bwise não encontrada.")

with col2:
    st.markdown('<h1 class="titulo-sistema">CONFERÊNCIA<br>ADIANTAMENTO SALARIAL</h1>', unsafe_allow_html=True)

with col3:
    if logo_macaneiro:
        st.image(logo_macaneiro, use_column_width=True)
    else:
        st.warning("Logo Maçaneiro não encontrada.")

st.markdown("---")

# -----------------------------------------------------------------------------
# INSTRUÇÕES DE USO (PASSO A PASSO KMM)
# -----------------------------------------------------------------------------
with st.expander("📖 Como extrair as planilhas do KMM (Passo a Passo)"):
    st.markdown("""
    ### 1. LISTA DE EVENTOS DE RECIBO DE PAGAMENTO
    **Caminho:** `Folha de Pagamento` ➔ `Folha de Pagamento` ➔ `Lista de Eventos de Recibos de Pagamento...`
    * **Competência Inicial:** Último adiantamento processado (mês anterior)
    * **Competência Final:** Adiantamento atual que está validando (mês atual)
    * **Tipo de Recibo:** `2 Adiantamento`
    * Clicar em **"Filtrar"** e salvar.

    ---
    ### 2. LISTA DE FUNCIONÁRIOS ATIVOS
    **Caminho:** `Folha de Pagamento` ➔ `Funcionários` ➔ `Registro...`
    * Clicar em **"Listar..."**
    * **Situação:** `Ativos`
    * Clicar em **"Filtrar"** e salvar.

    ---
    ### 3. LISTA DE PERÍODOS AQUISITIVOS E CONCESSIVOS (FÉRIAS)
    **Caminho:** `Folha de Pagamento` ➔ `Férias` ➔ `Lista de Períodos Aquisitivos e Concessivos...`
    * Clicar em **"Filtrar"** e salvar.
    """)

# -----------------------------------------------------------------------------
# FUNÇÕES DE TRATAMENTO DE DADOS E EXPORTAÇÃO
# -----------------------------------------------------------------------------
def limpar_moeda(valor):
    if pd.isna(valor): return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    v = str(valor).strip().replace('R$', '').strip()
    if not v: return 0.0
    if ',' in v and '.' in v:
        v = v.replace('.', '').replace(',', '.')
    elif ',' in v:
        v = v.replace(',', '.')
    try:
        return float(v)
    except ValueError:
        return 0.0

def limpar_matricula(val):
    if pd.isna(val): return ""
    v = str(val).strip()
    if v.endswith('.0'): v = v[:-2]
    return v

def verificar_optante(val):
    v = str(val).strip().lower()
    if v in ['0', '0.0', 'não', 'nao', 'n', 'false']:
        return False
    return True

def tem_direito_adiantamento(mes, ano, data_adm):
    if pd.isna(data_adm): return True
    data_ref_inicio = pd.Timestamp(year=ano, month=mes, day=1)
    if data_adm < data_ref_inicio: return True
    if data_adm.year == ano and data_adm.month == mes:
        return data_adm.day <= 6
    return False

def calcular_dias_ferias_no_mes(mes, ano, data_ini, data_fim):
    """Calcula os dias de férias que caem estritamente dentro de um mês (Mês Comercial de 30 dias)"""
    if pd.isna(data_ini) or pd.isna(data_fim): return 0
    
    primeiro_dia = pd.Timestamp(year=ano, month=mes, day=1)
    ultimo_dia_real = pd.Timestamp(year=ano, month=mes, day=1) + pd.offsets.MonthEnd(0)
    
    if data_fim < primeiro_dia or data_ini > ultimo_dia_real:
        return 0
        
    overlap_start = max(primeiro_dia, data_ini)
    overlap_end = min(ultimo_dia_real, data_fim)
    
    d_start = overlap_start.day
    if d_start == 31: d_start = 30
    
    d_end = overlap_end.day
    if d_end == 31: d_end = 30
    
    if overlap_start == primeiro_dia and overlap_end == ultimo_dia_real:
        return 30
        
    dias = d_end - d_start + 1
    return max(0, dias)

def formatar_moeda_br(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def gerar_excel_formatado(df, mes_ant, mes_atu):
    output = io.BytesIO()
    df_export = df.copy()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Conferência')
        workbook = writer.book
        worksheet = writer.sheets['Conferência']
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        
        for col_num in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        for row in range(2, worksheet.max_row + 1):
            for col_idx in [5, 6, 7, 8]:
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal='right')
            
            worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row, column=9).alignment = Alignment(horizontal='center')
            
            status_val = str(worksheet.cell(row=row, column=9).value)
            if 'Certo' in status_val:
                worksheet.cell(row=row, column=9).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            elif status_val == 'Errado':
                worksheet.cell(row=row, column=9).fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            elif status_val == 'Funcionário Novo':
                worksheet.cell(row=row, column=9).fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            else:
                worksheet.cell(row=row, column=9).fill = PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid')

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    val_str = str(cell.value)
                    if isinstance(cell.value, (int, float)): val_str += "R$ ,00"
                    if len(val_str) > max_len: max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    return output.getvalue()

@st.cache_data
def processar_dados(file_eventos, file_ativos, file_ferias):
    if file_eventos.name.endswith('.csv'): df_ev_raw = pd.read_csv(file_eventos, header=None, sep=None, engine='python')
    else: df_ev_raw = pd.read_excel(file_eventos, header=None)
        
    if file_ativos.name.endswith('.csv'): df_at_raw = pd.read_csv(file_ativos, header=None, sep=None, engine='python')
    else: df_at_raw = pd.read_excel(file_ativos, header=None)
        
    if file_ferias.name.endswith('.csv'): df_fe_raw = pd.read_csv(file_ferias, header=None, sep=None, engine='python')
    else: df_fe_raw = pd.read_excel(file_ferias, header=None)

    df_ev = df_ev_raw.iloc[2:, [1, 2, 11, 12, 13, 14, 16]].copy()
    df_ev.columns = ['Matricula', 'Nome_Ev', 'Mes', 'Ano', 'Cod_Evento', 'Nome_Evento', 'Valor_Provento']
    
    df_at = df_at_raw.iloc[2:, [1, 4, 44, 51, 62, 64]].copy()
    df_at.columns = ['Matricula', 'Nome', 'Data_Admissao', 'Categoria', 'Salario', 'Opta_Adiantamento']
    
    df_fe = df_fe_raw.iloc[2:, [4, 6, 14, 15]].copy()
    df_fe.columns = ['Matricula', 'Nome_Fe', 'Data_Ini_Ferias', 'Data_Fim_Ferias']

    df_ev['Matricula'] = df_ev['Matricula'].apply(limpar_matricula)
    df_at['Matricula'] = df_at['Matricula'].apply(limpar_matricula)
    df_fe['Matricula'] = df_fe['Matricula'].apply(limpar_matricula)
    
    df_ev = df_ev[df_ev['Matricula'] != ""]
    df_at = df_at[df_at['Matricula'] != ""]
    df_fe = df_fe[df_fe['Matricula'] != ""]

    df_ev['Valor_Provento'] = df_ev['Valor_Provento'].apply(limpar_moeda)
    df_at['Salario'] = df_at['Salario'].apply(limpar_moeda)
    df_ev['Mes'] = pd.to_numeric(df_ev['Mes'], errors='coerce').fillna(0).astype(int)
    df_ev['Ano'] = pd.to_numeric(df_ev['Ano'], errors='coerce').fillna(0).astype(int)
    df_ev['Cod_Evento'] = pd.to_numeric(df_ev['Cod_Evento'], errors='coerce').fillna(0).astype(int)
    df_at['Data_Admissao'] = pd.to_datetime(df_at['Data_Admissao'], errors='coerce', dayfirst=True)
    df_fe['Data_Ini_Ferias'] = pd.to_datetime(df_fe['Data_Ini_Ferias'], errors='coerce', dayfirst=True)
    df_fe['Data_Fim_Ferias'] = pd.to_datetime(df_fe['Data_Fim_Ferias'], errors='coerce', dayfirst=True)

    df_ev = df_ev[df_ev['Mes'] > 0]

    meses_disponiveis = sorted(df_ev['Mes'].unique())
    if len(meses_disponiveis) < 2:
        st.error("A planilha de EVENTOS não possui 2 meses distintos para comparação.")
        st.stop()
        
    mes_anterior = meses_disponiveis[0]
    mes_atual = meses_disponiveis[1]

    df_ev_100 = df_ev[df_ev['Cod_Evento'] == 100]
    total_ant_global = df_ev_100[df_ev_100['Mes'] == mes_anterior]['Valor_Provento'].sum()
    total_atu_global = df_ev_100[df_ev_100['Mes'] == mes_atual]['Valor_Provento'].sum()

    try: ano_ant = int(df_ev[df_ev['Mes'] == mes_anterior]['Ano'].mode()[0])
    except: ano_ant = 2026
    try: ano_atu = int(df_ev[df_ev['Mes'] == mes_atual]['Ano'].mode()[0])
    except: ano_atu = 2026

    matriculas_evento_invalido = df_ev[df_ev['Cod_Evento'] != 100]['Matricula'].unique()

    pivot_ev = df_ev_100.pivot_table(index='Matricula', columns='Mes', values='Valor_Provento', aggfunc='sum').reset_index()
    if mes_anterior not in pivot_ev.columns: pivot_ev[mes_anterior] = 0.0
    if mes_atual not in pivot_ev.columns: pivot_ev[mes_atual] = 0.0
    pivot_ev.rename(columns={mes_anterior: 'Valor_Mes_Anterior', mes_atual: 'Valor_Mes_Atual'}, inplace=True)

    report = pd.merge(df_at, pivot_ev, on='Matricula', how='left')
    report['Valor_Mes_Anterior'] = report['Valor_Mes_Anterior'].fillna(0.0)
    report['Valor_Mes_Atual'] = report['Valor_Mes_Atual'].fillna(0.0)

    resultados = []
    
    for _, row in report.iterrows():
        erros = []
        matricula = row['Matricula']
        categoria = str(row['Categoria']).strip()
        salario = row['Salario']
        val_ant = row['Valor_Mes_Anterior']
        val_atu = row['Valor_Mes_Atual']
        data_adm = row['Data_Admissao']
        
        is_aprendiz = "Aprendiz (Lei 10.097/2000)" in categoria
        optante = verificar_optante(row['Opta_Adiantamento'])
        direito_ant = tem_direito_adiantamento(mes_anterior, ano_ant, data_adm)
        direito_atu = tem_direito_adiantamento(mes_atual, ano_atu, data_adm)

        # CÁLCULO DE DIAS TRABALHADOS (DESCONTANDO FÉRIAS)
        ferias_func = df_fe[(df_fe['Matricula'] == matricula)]
        dias_fe_ant = 0
        dias_fe_atu = 0
        
        for _, v_row in ferias_func.iterrows():
            dias_fe_ant += calcular_dias_ferias_no_mes(mes_anterior, ano_ant, v_row['Data_Ini_Ferias'], v_row['Data_Fim_Ferias'])
            dias_fe_atu += calcular_dias_ferias_no_mes(mes_atual, ano_atu, v_row['Data_Ini_Ferias'], v_row['Data_Fim_Ferias'])
            
        if dias_fe_ant > 30: dias_fe_ant = 30
        if dias_fe_atu > 30: dias_fe_atu = 30
        
        dias_trab_ant = 30 - dias_fe_ant
        dias_trab_atu = 30 - dias_fe_atu

        # EXPECTATIVA DE VALOR (Regra: Menos de 15 dias trabalhados = Zero Adiantamento)
        if dias_trab_ant < 15:
            esperado_ant = 0.0
        else:
            esperado_ant = round((salario * 0.40) / 30 * dias_trab_ant, 2)
            
        if dias_trab_atu < 15:
            esperado_atu = 0.0
        else:
            esperado_atu = round((salario * 0.40) / 30 * dias_trab_atu, 2)

        if not optante:
            status_final = "Sem adiantamento (opcional)"
            descricao_erro = "Funcionário não optante (Coluna BM = 0)"
        elif is_aprendiz:
            if val_atu > 0 or val_ant > 0: erros.append("Aprendiz não deve receber adiantamento")
            status_final = "Errado" if erros else "Certo"
            descricao_erro = " - ".join(erros) if erros else "Correto (Aprendiz zerado)"
        else:
            if matricula in matriculas_evento_invalido: 
                erros.append("Contém evento diferente de 100")
                
            if not direito_atu:
                if val_atu > 0: erros.append("Recebeu indevidamente (Admitido após o dia 6)")
            else:
                # Se trabalhou 15 dias ou mais
                if dias_trab_atu >= 15:
                    if val_atu == 0: 
                        erros.append("Falta adiantamento no mês atual")
                    elif val_atu > 0 and abs(round(val_atu, 2) - esperado_atu) > 0.02:
                        if dias_trab_atu < 30:
                            erros.append(f"Cálculo de Férias incorreto (Esperado: R$ {esperado_atu:.2f} p/ {dias_trab_atu} dias trab.)")
                        else:
                            erros.append(f"Cálculo incorreto (Esperado: R$ {esperado_atu:.2f})")
                # Se trabalhou menos de 15 dias
                else:
                    if val_atu > 0:
                        erros.append(f"Recebeu indevidamente (Trabalhou apenas {dias_trab_atu} dias no mês)")

            # Divergência - Só avaliamos divergência cega se a expectativa para os dois meses era a mesma (ex: sem férias em nenhum, salário igual).
            if direito_ant and direito_atu:
                if esperado_ant == esperado_atu:
                    if round(val_ant, 2) != round(val_atu, 2): 
                        erros.append("Divergência de valor entre os meses")

            # Atribuição Final de Status
            if len(erros) > 0:
                status_final = "Errado"
                descricao_erro = " - ".join(erros)
            else:
                if not direito_atu:
                    status_final = "Não tem direito (admitido após dia 6)"
                    data_str = data_adm.strftime('%d/%m/%Y') if pd.notna(data_adm) else "Recente"
                    descricao_erro = f"Isento - Admitido em {data_str}"
                elif not direito_ant and direito_atu:
                    status_final = "Funcionário Novo"
                    descricao_erro = "Primeiro adiantamento (Isento de comp. c/ mês anterior)"
                elif dias_trab_atu < 15:
                    status_final = "Certo (Férias)"
                    descricao_erro = f"Isento de adiantamento ({dias_trab_atu} dias trab. no mês)"
                elif dias_trab_atu < 30:
                    status_final = "Certo (Férias)"
                    descricao_erro = f"Proporcional correto ({dias_trab_atu} dias trab. / {dias_fe_atu} dias férias)"
                else:
                    status_final = "Certo"
                    descricao_erro = "Sem divergências"
        
        data_adm_formatada = data_adm.strftime('%d/%m/%Y') if pd.notna(data_adm) else ""

        resultados.append({
            "Matricula": matricula,
            "Nome": row['Nome'],
            "Data de Admissão": data_adm_formatada,
            "Categoria": categoria,
            "Salario": salario,
            f"Adiantamento (Mês {mes_anterior})": val_ant,
            f"Adiantamento (Mês {mes_atual})": val_atu,
            "Diferença Entre Meses": val_atu - val_ant,
            "Status": status_final,
            "Motivo / Erros": descricao_erro
        })

    return pd.DataFrame(resultados), mes_anterior, mes_atual, total_ant_global, total_atu_global

# -----------------------------------------------------------------------------
# INTERFACE DO USUÁRIO
# -----------------------------------------------------------------------------
st.sidebar.header("📁 Importação de Dados")
file_eventos = st.sidebar.file_uploader("1. Upload: LISTA DE EVENTOS", type=["xlsx", "csv"])
file_ativos = st.sidebar.file_uploader("2. Upload: LISTA DE ATIVOS", type=["xlsx", "csv"])
file_ferias = st.sidebar.file_uploader("3. Upload: LISTA DE FÉRIAS", type=["xlsx", "csv"])

if file_eventos and file_ativos and file_ferias:
    with st.spinner("Processando dados, férias e aplicando regras de negócio..."):
        df_final, mes_ant, mes_atu, tot_ant_global, tot_atu_global = processar_dados(file_eventos, file_ativos, file_ferias)
    
    st.success(f"Dados processados com sucesso! Comparando Mês {mes_ant} x Mês {mes_atu}.")
    
    # --- FILTROS ---
    st.subheader("🔍 Filtros de Busca")
    col_f1, col_f2, col_f3 = st.columns(3)
    
    with col_f1: f_nome = st.text_input("Buscar por Nome")
    with col_f2: f_mat = st.text_input("Buscar por Matrícula")
    with col_f3:
        filtros_status = ["Todos", "Certo", "Certo (Férias)", "Errado", "Funcionário Novo", "Sem adiantamento (opcional)", "Não tem direito (admitido após dia 6)"]
        f_status = st.selectbox("Filtrar por Status", filtros_status)

    df_filtrado = df_final.copy()
    if f_nome: df_filtrado = df_filtrado[df_filtrado['Nome'].str.contains(f_nome, case=False, na=False)]
    if f_mat: df_filtrado = df_filtrado[df_filtrado['Matricula'].str.contains(f_mat, case=False, na=False)]
    if f_status != "Todos": df_filtrado = df_filtrado[df_filtrado['Status'] == f_status]

    # --- RESUMO GERAL ---
    st.markdown("---")
    st.subheader("📊 Resumo Geral")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Ativos", len(df_final))
    
    corretos = len(df_final[df_final['Status'].str.contains('Certo')])
    m2.metric("Corretos", corretos)
    
    m3.metric("Com Divergência", len(df_final[df_final['Status'] == 'Errado']))
    isentos = ['Funcionário Novo', 'Sem adiantamento (opcional)', 'Não tem direito (admitido após dia 6)']
    m4.metric("Isentos (Novos/Opcional)", len(df_final[df_final['Status'].isin(isentos)]))
    
    dif_total_global = tot_atu_global - tot_ant_global

    # --- TABELA DE RESULTADOS E DOWNLOAD ---
    st.markdown("---")
    st.markdown("### 📋 Tabela de Conferência (Por Empregado)")
    
    excel_ready = gerar_excel_formatado(df_filtrado, mes_ant, mes_atu)
    st.download_button(
        label="📥 Baixar Tabela de Conferência (Excel Formatado)",
        data=excel_ready,
        file_name=f"Conferencia_Adiantamentos_Mes_{mes_atu}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.write("")
    
    def colorir_status(val):
        if 'Certo' in val: return 'background-color: #d4edda; color: #155724'
        elif val == 'Errado': return 'background-color: #f8d7da; color: #721c24'
        elif val == 'Funcionário Novo': return 'background-color: #cce5ff; color: #004085'
        elif val == 'Não tem direito (admitido após dia 6)': return 'background-color: #fff3cd; color: #856404'
        else: return 'background-color: #e2e3e5; color: #383d41'
        
    st.dataframe(
        df_filtrado.style.map(colorir_status, subset=['Status'])\
                    .format({
                        "Salario": "R$ {:.2f}", 
                        f"Adiantamento (Mês {mes_ant})": "R$ {:.2f}", 
                        f"Adiantamento (Mês {mes_atu})": "R$ {:.2f}",
                        "Diferença Entre Meses": "R$ {:.2f}"
                    }),
        use_container_width=True,
        height=350
    )

    # --- ANÁLISES GRÁFICAS ---
    st.markdown("---")
    st.subheader("📈 Análises Gráficas da Empresa")
    
    st.markdown("<br><b>Comparativo de Adiantamentos (Mês a Mês)</b>", unsafe_allow_html=True)
    g1, g2 = st.columns(2)

    df_grafico = pd.DataFrame({
        "Mês": [f"Mês {mes_ant}", f"Mês {mes_atu}"],
        "Total Pago": [tot_ant_global, tot_atu_global]
    })
    df_grafico['Rótulo'] = df_grafico['Total Pago'].apply(formatar_moeda_br)

    with g1:
        base_chart = alt.Chart(df_grafico).encode(x=alt.X('Mês:N', sort=None))
        
        bars = base_chart.mark_bar(color="#1E3A8A").encode(
            y=alt.Y('Total Pago:Q', title="Total Pago (R$)"),
            tooltip=['Mês', 'Rótulo']
        )
        
        text_labels = base_chart.mark_text(
            align='center',
            baseline='bottom',
            dy=-5,
            fontWeight='bold',
            fontSize=13,
            color='#1E3A8A'
        ).encode(
            y=alt.Y('Total Pago:Q'),
            text='Rótulo:N'
        )
        
        st.altair_chart((bars + text_labels).properties(height=300), use_container_width=True)

    with g2:
        pie = alt.Chart(df_grafico).mark_arc(innerRadius=0, outerRadius=100).encode(
            theta=alt.Theta(field="Total Pago", type="quantitative"),
            color=alt.Color(field="Mês", type="nominal", scale=alt.Scale(range=['#1E3A8A', '#F59E0B'])),
            tooltip=['Mês', 'Rótulo']
        ).properties(height=300)
        st.altair_chart(pie, use_container_width=True)

    st.write("")
    st.markdown("<b>Diferença da Empresa (Mês Atual x Anterior)</b>", unsafe_allow_html=True)
    
    df_diff = pd.DataFrame({
        "Indicador": [f"Variação Mês {mes_ant} ➔ Mês {mes_atu}"],
        "Valor": [dif_total_global]
    })
    df_diff['Rótulo'] = df_diff['Valor'].apply(formatar_moeda_br)
    
    cor_diff = "#28a745" if dif_total_global >= 0 else "#dc3545"
    
    base_diff = alt.Chart(df_diff).encode(x=alt.X('Indicador:N'))
    
    bars_diff = base_diff.mark_bar(size=140, color=cor_diff).encode(
        y=alt.Y('Valor:Q', title="Diferença (R$)"),
        tooltip=['Indicador', 'Rótulo']
    )
    
    text_diff = base_diff.mark_text(
        align='center',
        baseline='bottom' if dif_total_global >= 0 else 'top',
        dy=-7 if dif_total_global >= 0 else 7,
        fontWeight='bold',
        fontSize=14,
        color=cor_diff
    ).encode(
        y=alt.Y('Valor:Q'),
        text='Rótulo:N'
    )
    
    st.altair_chart((bars_diff + text_diff).properties(height=300, width=400), use_container_width=False)

else:
    st.info("👈 Por favor, anexe as TRÊS planilhas no menu lateral para iniciar a conferência.")
